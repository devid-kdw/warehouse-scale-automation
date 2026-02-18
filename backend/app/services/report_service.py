"""Report services - statistics, inventory lists, and consumption tracking."""
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import List, Dict, Optional
from sqlalchemy import func, case

from ..extensions import db
from ..extensions import db
from ..models import Stock, Surplus, Transaction, Article, Batch, Location
from ..error_handling import AppError


def get_inventurna_lista(location_id: int) -> List[Dict]:
    """Get inventory count list (Article + Batch rows with current quantities)."""
    stock_qty_col = func.coalesce(Stock.quantity, 0).label('stock_qty')
    surplus_qty_col = func.coalesce(Surplus.quantity, 0).label('surplus_qty')
    
    query = db.session.query(
        Article, Batch,
        stock_qty_col,
        surplus_qty_col
    ).join(
        Batch, Batch.article_id == Article.id
    ).outerjoin(
        Stock, (Stock.batch_id == Batch.id) & (Stock.location_id == location_id)
    ).outerjoin(
        Surplus, (Surplus.batch_id == Batch.id) & (Surplus.location_id == location_id)
    ).filter(
        Article.is_active == True
    ).order_by(Article.article_no, Batch.batch_code)
    
    results = query.all()
    items = []
    for art, batch, s_qty, sur_qty in results:
        # DB quantities are now authoritative units
        s_val = float(s_qty)
        sur_val = float(sur_qty)
        total_val = s_val + sur_val
        
        items.append({
            'article_id': art.id,
            'article_no': art.article_no,
            'description': art.description,
            'batch_id': batch.id,
            'batch_code': batch.batch_code,
            'stock': s_val,
            'surplus': sur_val,
            'total': total_val,
            'uom': art.uom
        })
    return items


def get_surplus_lista(location_id: int) -> List[Dict]:
    """Get detailed Surplus report."""
    query = db.session.query(
        Surplus, Article, Batch
    ).join(
        Article, Surplus.article_id == Article.id
    ).join(
        Batch, Surplus.batch_id == Batch.id
    ).filter(
        Surplus.location_id == location_id,
        Surplus.quantity > 0 # Unit-aware check
    ).order_by(Article.article_no)
    
    results = query.all()
    return [{
        'article_no': art.article_no,
        'description': art.description,
        'batch_code': batch.batch_code,
        'quantity': float(s.quantity),
        'uom': art.uom,
        'updated_at': s.updated_at.isoformat() if s.updated_at else None
    } for s, art, batch in results]


def get_consumption_stats(days: int = 30) -> List[Dict]:
    """Get consumption stats per article over a period."""
    since = datetime.now(timezone.utc) - timedelta(days=days)
    
    consumption_types = [Transaction.TX_STOCK_CONSUMED, Transaction.TX_SURPLUS_CONSUMED]
    
    query = db.session.query(
        Article.article_no,
        Article.description,
        Article.uom,
        func.sum(Transaction.quantity).label('total_qty'),
        func.count(Transaction.id).label('hit_count')
    ).join(
        Transaction, Transaction.article_id == Article.id
    ).filter(
        Transaction.tx_type.in_(consumption_types),
        Transaction.occurred_at >= since
    ).group_by(Article.id).order_by(func.sum(Transaction.quantity).asc()) 
    # ASC because quantities are negative. Most negative (largest consumption) first.
    
    results = query.all()
    return [{
        'article_no': r.article_no,
        'description': r.description,
        'uom': r.uom,
        'quantity': float(r.total_qty or 0),
        'hit_count': r.hit_count
    } for r in results]


def get_reorder_risk_lista(location_id: int, state_filter: str = 'active', include_green: bool = False) -> List[Dict]:
    """Get articles at or near reorder threshold (including 10% yellow zone)."""
    
    # Calculate current total stock (unit-aware)
    stock_subq = db.session.query(
        Stock.article_id,
        func.sum(Stock.quantity).label('total_stock')
    ).filter(Stock.location_id == location_id).group_by(Stock.article_id).subquery()
    
    query = db.session.query(
        Article,
        func.coalesce(stock_subq.c.total_stock, 0).label('stock_qty')
    ).outerjoin(
        stock_subq, Article.id == stock_subq.c.article_id
    )
    
    if state_filter == 'active':
        query = query.filter(Article.is_active == True)
    elif state_filter == 'inactive':
        query = query.filter(Article.is_active == False)
        
    query = query.filter(Article.reorder_threshold > 0)
    
    results = query.all()
    items = []
    for art, s_qty in results:
        threshold = Decimal(str(art.reorder_threshold))
        current = Decimal(str(s_qty))
        
        # Risk Logic:
        # RED: <= threshold
        # YELLOW: > threshold AND <= threshold * 1.1 (within 10% above)
        # GREEN: > threshold * 1.1
        
        risk_level = 'GREEN'
        if current <= threshold:
            risk_level = 'RED'
        elif current <= threshold * Decimal('1.1'):
            risk_level = 'YELLOW'
            
        if risk_level != 'GREEN' or include_green:
            items.append({
                'article_no': art.article_no,
                'description': art.description,
                'uom': art.uom,
                'stock': float(current), # Unit-aware
                'threshold': float(threshold),
                'risk_level': risk_level
            })
            
    # Sort by risk (RED > YELLOW > GREEN) then by closeness to threshold
    level_order = {'RED': 0, 'YELLOW': 1, 'GREEN': 2}
    return sorted(items, key=lambda x: (level_order.get(x['risk_level'], 3), x['stock'] / x['threshold'] if x['threshold'] > 0 else 0))


def get_top_20_monthly_consumers() -> List[Dict]:
    """Get top 20 consumers (Article hits) in the last 30 days."""
    since = datetime.now(timezone.utc) - timedelta(days=30)
    consumption_types = [Transaction.TX_STOCK_CONSUMED, Transaction.TX_SURPLUS_CONSUMED]
    
    query = db.session.query(
        Article.article_no,
        Article.description,
        Article.uom,
        func.sum(Transaction.quantity).label('total_qty'),
        func.count(Transaction.id).label('hit_count')
    ).join(
        Transaction, Transaction.article_id == Article.id
    ).filter(
        Transaction.tx_type.in_(consumption_types),
        Transaction.occurred_at >= since
    ).group_by(Article.id).order_by(func.sum(Transaction.quantity).asc()).limit(20) 
    # ASC because quantities are negative. Most negative (largest consumption) first.
    # Logic verified: -100 < -10, so ASC gives -100 first, which is correct for "Top Consumer".
    
    results = query.all()
    return [{
        'article_no': r.article_no,
        'description': r.description,
        'uom': r.uom,
        'quantity': float(r.total_qty or 0),
        'hit_count': r.hit_count
    } for r in results]


import io
# openpyxl and fpdf2 moved to lazy imports inside functions for startup safety

def export_inventurna_to_excel(location_id: int) -> io.BytesIO:
    """Export inventory count list to formatted Excel (Lazy Import)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise AppError(
            'FAILED_DEPENDENCY', 
            'Excel export requires openpyxl. Please contact support or check server logs.'
        )

    items = get_inventurna_lista(location_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventurna Lista"
    
    headers = ["Article No", "Description", "Batch code", "Stock", "Surplus", "Total", "UOM"]
    ws.append(headers)
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for item in items:
        ws.append([
            item['article_no'], item['description'], item['batch_code'],
            item['stock'], item['surplus'], item['total'], item['uom']
        ])
        
    # Column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    # Freeze panes
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_surplus_to_excel(location_id: int) -> io.BytesIO:
    """Export surplus list to formatted Excel (Lazy Import)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise AppError('FAILED_DEPENDENCY', 'Excel export requires openpyxl.')

    items = get_surplus_lista(location_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Surplus Report"
    
    headers = ["Article No", "Description", "Batch code", "Surplus", "UOM", "Last Updated"]
    ws.append(headers)
    
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    for cell in ws[1]:
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center")
        
    for item in items:
        ws.append([
            item['article_no'], item['description'], item['batch_code'],
            item['quantity'], item['uom'], item['updated_at']
        ])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ReportPDF moved inside lazy import scope or made independent if possible
# Since it inherits FPDF, it must be defined where FPDF is available or used inside function

def export_inventurna_to_pdf(location_id: int) -> bytes:
    """Export inventory count list to clean PDF (Lazy Import)."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise AppError(
            'FAILED_DEPENDENCY', 
            'PDF export requires fpdf2. Please contact support or check server logs.'
        )

    class ReportPDF(FPDF):
        def header(self):
            self.set_font('helvetica', 'B', 15)
            self.cell(0, 10, 'Paint Manager - Report', border=False, align='C', ln=1)
            self.set_font('helvetica', 'I', 10)
            self.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=1, align='C')
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font('helvetica', 'I', 8)
            self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')

    items = get_inventurna_lista(location_id)
    pdf = ReportPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 10, "Inventurna Lista", ln=1)
    
    # Table header
    pdf.set_font('helvetica', 'B', 9)
    cols = [("Article No", 40), ("Batch", 30), ("Stock", 25), ("Surplus", 25), ("Total", 25), ("UOM", 15)]
    for txt, w in cols:
        pdf.cell(w, 8, txt, 1, 0, 'C')
    pdf.ln()
    
    # Data
    pdf.set_font('helvetica', '', 8)
    for item in items:
        # Check for page break
        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 9)
            for txt, w in cols:
                pdf.cell(w, 8, txt, 1, 0, 'C')
            pdf.ln()
            pdf.set_font('helvetica', '', 8)
            
        pdf.cell(40, 7, str(item['article_no']), 1)
        pdf.cell(30, 7, str(item['batch_code']), 1)
        pdf.cell(25, 7, f"{item['stock']:.2f}", 1, 0, 'R')
        pdf.cell(25, 7, f"{item['surplus']:.2f}", 1, 0, 'R')
        pdf.cell(25, 7, f"{item['total']:.2f}", 1, 0, 'R')
        pdf.cell(15, 7, str(item['uom']), 1, 0, 'C')
        pdf.ln()
        
    return pdf.output()


def export_surplus_to_pdf(location_id: int) -> bytes:
    """Export surplus list to clean PDF (Lazy Import)."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise AppError('FAILED_DEPENDENCY', 'PDF export requires fpdf2.')

    class ReportPDF(FPDF):
        def header(self):
            self.set_font('helvetica', 'B', 15)
            self.cell(0, 10, 'Paint Manager - Surplus Report', border=False, align='C', ln=1)
            self.set_font('helvetica', 'I', 10)
            self.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=1, align='C')
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font('helvetica', 'I', 8)
            self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')

    items = get_surplus_lista(location_id)
    pdf = ReportPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 10, f"Surplus Stock (Location {location_id})", ln=1)
    
    # Table header
    pdf.set_font('helvetica', 'B', 9)
    # Article No (40), Batch (30), Surplus (30), UOM (15), Desc (Rest)
    cols = [("Article No", 40), ("Batch", 30), ("Surplus", 30), ("UOM", 15), ("Description", 70)]
    for txt, w in cols:
        pdf.cell(w, 8, txt, 1, 0, 'C')
    pdf.ln()
    
    # Data
    pdf.set_font('helvetica', '', 8)
    for item in items:
        # Check for page break
        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 9)
            for txt, w in cols:
                pdf.cell(w, 8, txt, 1, 0, 'C')
            pdf.ln()
            pdf.set_font('helvetica', '', 8)
            
        pdf.cell(40, 7, str(item['article_no']), 1)
        pdf.cell(30, 7, str(item['batch_code']), 1)
        pdf.cell(30, 7, f"{item['quantity']:.2f}", 1, 0, 'R')
        pdf.cell(15, 7, str(item['uom']), 1, 0, 'C')
        # Truncate description if too long
        desc = str(item['description'] or '')[:40]
        pdf.cell(70, 7, desc, 1, 0, 'L')
        pdf.ln()
        
    return pdf.output()


def get_reporting_stats() -> Dict:
    """Get statistics for missing article reports."""
    from ..models import MissingArticleReport
    
    total = MissingArticleReport.query.count()
    by_status = db.session.query(
        MissingArticleReport.status, 
        func.count(MissingArticleReport.id)
    ).group_by(MissingArticleReport.status).all()
    
    return {
        'total_count': total,
        'status_breakdown': dict(by_status),
        'generated_at': datetime.now(timezone.utc).isoformat()
    }
